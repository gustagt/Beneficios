from PIL import Image, ImageDraw
from flask import render_template, send_file

import openpyxl
from win32com import client
import win32com.client
import os
from pathlib import Path
import time

from datetime import date, timedelta
from models import erros as e
from models.beneficiario import Beneficiario as bn
from models.protocolo import Protocolo as pt
from models.taxas import Taxas as tx
from models.deficiente import Deficiente as df
from models.idoso import Idoso as id

dataAtual = date.today()
dataVencimento = dataAtual + timedelta(days=2)
idProtudo = '8'
idSegmento = '5'
idvalor = '6'
idOrgao = '0752'
codigoBeneficio = '10'
path = Path().absolute()


class SolicitarBoletoController:
    def __init__(self):
        None

    def solicitarBoletoGET(request):

        cpf = request.cookies.get('cpf')
        nCredencial = request.cookies.get('nCredencial')
        tpCredencial = request.cookies.get('tpCredencial')

        if tpCredencial.lower() == 'deficiente':
            credencial = df.selectByCpf(cpf)
            if not (credencial.empty):
                nCredencialResp = credencial['n_credencial'].values[0]
                beneficiario = bn.selectByCpf(cpf)
                if nCredencialResp == int(nCredencial):
                    protocolo = pt(
                        beneficiario.cpf, 'Não pago', "Segunda via - Boleto").add()
                    SolicitarBoletoController.creatCodigoBarra(
                        beneficiario, protocolo)
                    file = path / 'static' / 'boletos' / \
                        '{0}.pdf'.format(protocolo.nProtocolo)
                    return send_file(file), 200

        elif tpCredencial.lower() == 'idoso':
            credencial = id.selectByCpf(cpf)
            if not (credencial.empty):
                nCredencialResp = credencial['n_credencial'].values[0]
                beneficiario = bn.selectByCpf(cpf)
                if nCredencialResp == int(nCredencial):
                    protocolo = pt(
                        beneficiario.cpf, 'Não pago', "Segunda via - Boleto").add()
                    SolicitarBoletoController.creatCodigoBarra(
                        beneficiario, protocolo)
                    file = path / 'static' / 'boletos' / \
                        '{0}.pdf'.format(protocolo.nProtocolo)
                    return send_file(file), 200

        return render_template('cliente/erro.html', mensagem=e.Erros.consulta.value), 400

    def imprimirBoleto(beneficiario: bn, protocolo: pt, codigo44):

        # Open Microsoft Excel
        win32com.client.pythoncom.CoInitialize()
        excel = client.Dispatch("Excel.Application")
        # abre arquivo xlsx
        wrkb = openpyxl.load_workbook(
            path / 'documentos' / 'templatesDocumentos' / 'entrada.xlsx')
        ws = wrkb.worksheets[0]

        # via cliente excel
        mensagem = ws['A6']
        nomePessoa = ws['A14']
        dataValidade = ws['A16']
        dataEmissao = ws['C16']
        exercicio = ws['E16']
        nIdentificacao = ws['H15']
        endereco = ws['A18']
        nEndereco = ws['B20']
        bairro = ws['B21']
        cep = ws['B22']
        valor = ws['H17']
        vencimento = ws['H18']
        valorBoleto = ws['H21']

        # via banco excel
        nomePessoaB = ws['A33']
        dataValidadeB = ws['A35']
        dataEmissaoB = ws['C35']
        exercicioB = ws['E35']
        nIdentificacaoB = ws['H34']
        enderecoB = ws['A37']
        nEnderecoB = ws['B39']
        bairroB = ws['B40']
        cepB = ws['B41']
        valorB = ws['H36']
        vencimentoB = ws['H37']
        valorBoletoB = ws['H40']

        # linha Digitavel excel
        campo1 = ws['A45']
        campo2 = ws['C45']
        campo3 = ws['E45']
        campo4 = ws['G45']

        # acessar valores
        mensagem.value = e.Mensagens.mensagemBoleto.value

        # via cliente
        nomePessoa.value = beneficiario.nome
        dataValidade.value = dataVencimento
        dataEmissao.value = dataAtual
        exercicio.value = dataAtual.year
        nIdentificacao.value = protocolo.nProtocolo
        endereco.value = beneficiario.rua
        nEndereco.value = beneficiario.numComplemento
        bairro.value = beneficiario.bairro
        cep.value = beneficiario.cep
        valor.value = str(tx.selectTaxa()).replace(".", ",")
        vencimento.value = dataVencimento
        valorBoleto.value = str(tx.selectTaxa()).replace(".", ",")

        # via Banco
        nomePessoaB.value = beneficiario.nome
        dataValidadeB.value = dataVencimento
        dataEmissaoB.value = dataAtual
        exercicioB.value = dataAtual.year
        nIdentificacaoB.value = protocolo.nProtocolo
        enderecoB.value = beneficiario.rua + ' - ' + beneficiario.cidade
        nEnderecoB.value = beneficiario.numComplemento
        bairroB.value = beneficiario.bairro
        cepB.value = beneficiario.cep
        valorB.value = str(tx.selectTaxa()).replace(".", ",")
        vencimentoB.value = dataVencimento
        valorBoletoB.value = str(tx.selectTaxa()).replace(".", ",")

        # linha Digitavel excel
        campo1.value = str(codigo44)[
            :11] + ' ' + SolicitarBoletoController.digitoVerificadorM10(str(codigo44)[:11])
        campo2.value = str(codigo44)[
            11:22] + ' ' + SolicitarBoletoController.digitoVerificadorM10(str(codigo44)[11:22])
        campo3.value = str(codigo44)[
            22:33] + ' ' + SolicitarBoletoController.digitoVerificadorM10(str(codigo44)[22:33])
        campo4.value = str(codigo44)[
            33:44] + ' ' + SolicitarBoletoController.digitoVerificadorM10(str(codigo44)[33:44])

        # inserindo o codigo de barras
        img = openpyxl.drawing.image.Image(
            path / 'documentos' / 'templatesDocumentos' / '{0}.png'.format(beneficiario.cpf))
        img.anchor = 'A47'
        ws.add_image(img)

        # salva arquivo xlsx
        wrkb.save(path / 'documentos' / 'templatesDocumentos' /
                  '{0}.xlsx'.format(beneficiario.cpf))

        # Read Excel File
        sheets = excel.Workbooks.Open(
            path / 'documentos' / 'templatesDocumentos' / '{0}.xlsx'.format(beneficiario.cpf))

        teste = path / 'static' / 'boletos' / \
            '{0}.pdf'.format(protocolo.nProtocolo)
        # Convert into PDF File
        sheets.Worksheets[0].ExportAsFixedFormat(0, teste.__str__())

        # fecha o excel
        excel.Quit()

        # delay para fechar e excluir os arquivos
        time.sleep(1)
        if os.path.exists(path / 'documentos' / 'templatesDocumentos' / '{0}.xlsx'.format(beneficiario.cpf)):
            os.remove(path / 'documentos' / 'templatesDocumentos' /
                      '{0}.xlsx'.format(beneficiario.cpf))

        if os.path.exists(path / 'documentos' / 'templatesDocumentos' / '{0}.png'.format(beneficiario.cpf)):
            os.remove(path / 'documentos' / 'templatesDocumentos' /
                      '{0}.png'.format(beneficiario.cpf))

    def codBarraImage(valor, beneficiario: bn, posX=150, posY=0, height=60):

        # padrão 2 por 5 intercalado ( utilizado em boletos bancários )
        padrao = ('00110', '10001', '01001', '11000', '00101',
                  '10100', '01100', '00011', '10010', '01010')

        # criando imagem
        imagem = Image.new('RGB', (570, 60), 'white')
        draw = ImageDraw.Draw(imagem)

        # verificando se o conteudo para gerar barra é impar, se for,
        # adiciona 0 no inicial para fazer intercalação em seguida dos pares

        if (len(valor) % 2) != 0:
            valor = '0' + valor

        # faz intercalação dos pares
        l = ''
        for i in range(0, len(valor), 2):
            p1 = padrao[int(valor[i])]
            p2 = padrao[int(valor[i+1])]
            for p in range(0, 5):
                l += p1[:1] + p2[:1]
                p1 = p1[1:]
                p2 = p2[1:]

        # gerando espaços e barras
        barra = True
        b = ''

        # P = preto
        # B = banco

        for i in range(0, len(l)):
            if l[i] == '0':
                if barra:
                    b += 'P'
                    barra = False
                else:
                    b += 'B'
                    barra = True
            else:
                if barra:
                    b += 'PPP'
                    barra = False
                else:
                    b += 'BBB'
                    barra = True

        # concatena inicio e fim
        b = 'PBPB' + b + 'PPPBP'

        # P = preto
        # B = banco

        # percorre toda a string b e onde for P pinta de preto, onde for B pinta de banco

        for i in range(0, len(b)):
            if b[i] == 'P':
                draw.line((posX, posY, posX, posY + height), 'black')
            else:
                draw.line((posX, posY, posX, posY + height), 'white')
            posX += 1

        imagem.save(path / 'documentos' / 'templatesDocumentos' /
                    '{0}.png'.format(str(beneficiario.cpf)))
        return

    def digitoVerificadorM10(codigo):
        soma = 0
        count = 0
        for i in codigo[::-1]:
            if (count % 2 == 0):
                valida = int(i) * 2
                if (valida >= 10):
                    char1 = str(valida)[0]
                    char2 = str(valida)[1]
                    soma += int(char1) + int(char2)
                else:
                    soma += int(i) * 2
            else:
                soma += int(i)
            count += 1

        dac = 10 - (soma % 10)
        return str(dac)

    def creatCodigoBarra(beneficiario: bn, protocolo: pt):

        valor = str(tx.selectTaxa()).replace(".", "").zfill(11)

        campoLivre = str(dataVencimento).replace("-", "") + codigoBeneficio + \
            str(protocolo.nProtocolo).zfill(7) + \
            str(dataAtual).replace("-", "")

        codigoPre = idProtudo + idSegmento + idvalor + valor + idOrgao + campoLivre

        digitoVerificadorGeral = SolicitarBoletoController.digitoVerificadorM10(
            codigoPre)

        codigo44 = idProtudo + idSegmento + idvalor + \
            digitoVerificadorGeral + valor + idOrgao + campoLivre

        SolicitarBoletoController.codBarraImage(codigo44, beneficiario)
        SolicitarBoletoController.imprimirBoleto(
            beneficiario, protocolo, codigo44)
