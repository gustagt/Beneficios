import sys
from PIL import Image, ImageDraw, ImageFont

import openpyxl  
from win32com import client
import os 
import time

from datetime import date, timedelta


def imprimirBoleto(beneficiario):
        
    # Open Microsoft Excel
    excel = client.Dispatch("Excel.Application")
    # abre arquivo xlsx
    wrkb = openpyxl.load_workbook('C:\\Users\\0103250\\Documents\\Beneficios\\Api\\documentos\\templatesDocumentos\\entrada.xlsx') 
    ws = wrkb.worksheets[0]

    # via cliente excel
    mensagem = ws['A6']
    nomePessoa = ws['A14']
    dataValidade = ws['A16']
    dataEmissao = ws['C16']
    exercicio = ws['E16']
    nIdentificacao = ws['H15']
    endereco = ws['A18']
    nEndereco = ws['B20']
    bairro = ws['B21']
    cep = ws['B22']
    valor = ws['H17']
    vencimento = ws['H18']
    valorBoleto = ws['H21']

    # via banco excel
    nomePessoaB = ws['A33']
    dataValidadeB = ws['A35']
    dataEmissaoB = ws['C35']
    exercicioB = ws['E35']
    nIdentificacaoB = ws['H34']
    enderecoB = ws['A37']
    nEnderecoB = ws['B39']
    bairroB = ws['B40']
    cepB = ws['B41']
    valorB = ws['H36']
    vencimentoB = ws['H37']

    # linha Digitavel excel
    campo1 = ws['A45']
    campo2 = ws['C45']
    campo3 = ws['E45']
    campo4 = ws['G45']

    # acessar valores
    nomePessoa.value = 'gustavo augusto'

    # inserindo o codigo de barras
    img = openpyxl.drawing.image.Image('C:\\Users\\0103250\\Documents\\Beneficios\\Api\\documentos\\templatesDocumentos\\{0}.png'.format(beneficiario)) 
    img.anchor = 'A47'
    ws.add_image(img)

    # salva arquivo xlsx
    wrkb.save('C:\\Users\\0103250\\Documents\\Beneficios\\Api\\documentos\\templatesDocumentos\\{0}.xlsx'.format(beneficiario))

    # Read Excel File
    sheets = excel.Workbooks.Open('C:\\Users\\0103250\\Documents\\Beneficios\\Api\\documentos\\templatesDocumentos\\{0}.xlsx'.format(beneficiario))

    
    # Convert into PDF File
    sheets.Worksheets[0].ExportAsFixedFormat(0, 'C:\\Users\\0103250\\Documents\\Beneficios\\Api\\documentos\\templatesDocumentos\\{0}.pdf'.format(beneficiario))

    # fecha o excel
    excel.Quit()

    # delay para fechar e excluir os arquivos
    time.sleep(1)
    if os.path.exists('C:\\Users\\0103250\\Documents\\Beneficios\\Api\\documentos\\templatesDocumentos\\{0}.xlsx'.format(beneficiario)): 
        os.remove('C:\\Users\\0103250\\Documents\\Beneficios\\Api\\documentos\\templatesDocumentos\\{0}.xlsx'.format(beneficiario)) 
        
    if os.path.exists('C:\\Users\\0103250\\Documents\\Beneficios\\Api\\documentos\\templatesDocumentos\\{0}.png'.format(beneficiario)): 
        os.remove('C:\\Users\\0103250\\Documents\\Beneficios\\Api\\documentos\\templatesDocumentos\\{0}.png'.format(beneficiario)) 
            
            
def codBarraImage(valor, cpf, posX=150, posY=0, height = 60):

    # padrão 2 por 5 intercalado ( utilizado em boletos bancários )
    padrao = ('00110', '10001', '01001', '11000', '00101',
          '10100', '01100', '00011', '10010', '01010')

    # criando imagem
    imagem = Image.new('RGB', (570,60), 'white')
    draw = ImageDraw.Draw(imagem)

    # verificando se o conteudo para gerar barra é impar, se for,
    # adiciona 0 no inicial para fazer intercalação em seguida dos pares 

    if (len(valor) % 2) != 0:
      valor= '0' + valor

    # faz intercalação dos pares
    l=''
    for i in range(0,len(valor),2):
      p1=padrao[int(valor[i])]
      p2=padrao[int(valor[i+1])]
      for p in range(0,5):
        l+=p1[:1] + p2[:1]
        p1=p1[1:]
        p2=p2[1:]

    # gerando espaços e barras
    barra=True
    b=''

    # P = preto
    # B = banco

    for i in range(0,len(l)):
      if l[i] == '0':
        if barra:
          b+='P'
          barra=False
        else:
          b+='B'
          barra=True
      else:
        if barra:
          b+='PPP'
          barra=False
        else:
          b+='BBB'
          barra=True

    # concatena inicio e fim
    b='PBPB' + b + 'PPPBP'

    # P = preto
    # B = banco 

    # percorre toda a string b e onde for P pinta de preto, onde for B pinta de banco 

    for i in range(0,len(b)):
      if b[i] == 'P':
        draw.line((posX,posY,posX,posY + height),'black')
      else:
        draw.line((posX,posY,posX,posY + height),'white')
      posX+=1
    
    imagem.save('C:\\Users\\0103250\\Documents\\Beneficios\\Api\\documentos\\templatesDocumentos\\%s.png'%(str(cpf)))
    return 



def digitoVerificadorM10(codigo):
    soma = 0
    count = 0
    for i in codigo[::-1]:
        if (count % 2 == 0):
            valida = int(i) * 2
            if (valida >= 10): 
                char1 = str(valida)[0]
                char2 = str(valida)[1]
                soma += int(char1) + int(char2)
            else:
                soma += int(i) * 2
        else: 
            soma += int(i) 
        count += 1
    
    dac = 10 - (soma%10)
    return str(dac)

def creatCodigoBarra(nProcotolo , beneficiario):
    idProtudo = '8'
    idSegmento = '5'
    idvalor = '8'

    valor = '158,76'
    valor = valor.replace(',','').zfill(11)
    idOrgao = '0752'
    
    dataAtual = date.today()
    dataVencimento = dataAtual + timedelta(days=2)
    
    codigoBeneficio = '10'
    
    campoLivre = str(dataVencimento).replace("-","") + codigoBeneficio + str(nProcotolo).zfill(7)+ str(dataAtual).replace("-","")
    
    codigoPre = idProtudo + idSegmento + idvalor + valor + idOrgao + campoLivre
    
    digitoVerificadorGeral = digitoVerificadorM10(codigoPre)

    codigo44 = idProtudo + idSegmento + idvalor + digitoVerificadorGeral + valor + idOrgao + campoLivre
    
    codBarraImage(codigo44, beneficiario)
    imprimirBoleto(beneficiario)

creatCodigoBarra('123', 123)

    
    # zfill



    


